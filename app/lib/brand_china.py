import pandas as pd
from transform.path import UPLOAD_CHINA_FORM
from openpyxl import load_workbook


def excel2df(file_path):
    try:
        df = pd.read_excel(file_path)
        df = df.fillna("")
        # 데이터타입 int -> str
        df = df.applymap(lambda x: str(x))
        return df
    except:
        print("엑셀파일을 불러오는데 실패했습니다.")
        return 0


def generate_df(brand, china_columns):
    # 빈 데이터프레임 선언
    data = pd.DataFrame()

    # 세부사이즈 하위컬럼명 변경
    brand.rename(columns={'Unnamed: 8': '세부사이즈(상의)2', 'Unnamed: 9': '세부사이즈(상의)3', 'Unnamed: 10': '세부사이즈(상의)4',
                          'Unnamed: 12': '세부사이즈(하의)2', 'Unnamed: 13': '세부사이즈(하의)3', 'Unnamed: 14': '세부사이즈(하의)4',
                          'Unnamed: 15': '세부사이즈(하의)5', 'Unnamed: 16': '세부사이즈(하의)6'}, inplace=True)

    for brand_row_num in range(1, len(brand)):

        if brand.iloc[brand_row_num]["색상"] == "":
            list_col = ["ONE COLOR"]
        else:
            list_col = brand.iloc[brand_row_num]["색상"].split(',')

        if brand.iloc[brand_row_num]["사이즈"] == "":
            list_size = ["FREE"]
        else:
            list_size = brand.iloc[brand_row_num]["사이즈"].split(',')

        list_stock = brand.iloc[brand_row_num]["재고수량"]
        list_stock = list_stock.split(',')
        list_stock = [i.strip() for i in list_stock]
        list_stock = list(map(int, list_stock))
        

            # 상의 세부사이즈만 있을 때
        if brand.iloc[brand_row_num]["세부사이즈(상의)"] != "" and brand.iloc[brand_row_num]["세부사이즈(하의)"] == "":
            list_size_u1 = brand.iloc[brand_row_num]["세부사이즈(상의)"].split(',')
            list_size_u2 = brand.iloc[brand_row_num]["세부사이즈(상의)2"].split(',')
            list_size_u3 = brand.iloc[brand_row_num]["세부사이즈(상의)3"].split(',')
            list_size_u4 = brand.iloc[brand_row_num]["세부사이즈(상의)4"].split(',')
        # 하의 세부사이즈만 있을 때
        elif brand.iloc[brand_row_num]["세부사이즈(하의)"] != "" and brand.iloc[brand_row_num]["세부사이즈(상의)"] == "":
            list_size_d1 = brand.iloc[brand_row_num]["세부사이즈(하의)"].split(',')
            list_size_d2 = brand.iloc[brand_row_num]["세부사이즈(하의)2"].split(',')
            list_size_d3 = brand.iloc[brand_row_num]["세부사이즈(하의)3"].split(',')
            list_size_d4 = brand.iloc[brand_row_num]["세부사이즈(하의)4"].split(',')
            list_size_d5 = brand.iloc[brand_row_num]["세부사이즈(하의)5"].split(',')
            list_size_d6 = brand.iloc[brand_row_num]["세부사이즈(하의)6"].split(',')
        # 모두 공백일 때
        else:
            list_size_u1 = ""
            list_size_u2 = ""
            list_size_u3 = ""
            list_size_u4 = ""
            list_size_d1 = ""
            list_size_d2 = ""
            list_size_d3 = ""
            list_size_d4 = ""
            list_size_d5 = ""
            list_size_d6 = ""
        # 값 할당
        for nol_col in range(len(list_col)):
            for nol_size in range(len(list_size)):
                # 세부사이즈(상의)만 있을 때
                if len(list_col) >= len(list_size):
                    k = nol_col
                else:
                    k = nol_size
                if brand.iloc[brand_row_num]["세부사이즈(상의)"] != "":
                    # 세부사이즈(상의) 개수 == 사이즈 개수
                    if len(list_size_u1) == len(list_size):
                        val = {'상의-어깨너비': list_size_u1[nol_size],
                               '상의-가슴너비': list_size_u2[nol_size],
                               '상의-소매길이': list_size_u3[nol_size],
                               '상의-총장(앞)': list_size_u4[nol_size]
                               }
                    else:
                        val = {'상의-어깨너비': list_size_u1,
                               '상의-가슴너비': list_size_u2,
                               '상의-소매길이': list_size_u3,
                               '상의-총장(앞)': list_size_u4
                               }
                # 세부사이즈(하의)만 있을 때
                elif brand.iloc[brand_row_num]["세부사이즈(하의)"] != "":
                    # 세부사이즈(하의) 개수 == 사이즈 개수
                    if len(list_size_d1) == len(list_size):
                        val = {'하의-총장(아웃심)': list_size_d1[nol_size],
                               '하의-허리': list_size_d2[nol_size],
                               '하의-엉덩이': list_size_d3[nol_size],
                               '하의-허벅지': list_size_d4[nol_size],
                               '하의-밑위': list_size_d5[nol_size],
                               '하의-밑단': list_size_d6[nol_size]
                               }
                    else:
                        val = {'하의-총장(아웃심)': list_size_d1,
                               '하의-허리': list_size_d2,
                               '하의-엉덩이': list_size_d3,
                               '하의-허벅지': list_size_d4,
                               '하의-밑위': list_size_d5,
                               '하의-밑단': list_size_d6
                               }
                # 세부사이즈가 모두 공백일때
                else:
                    val = {'상의-어깨너비': list_size_u1,
                           '상의-가슴너비': list_size_u2,
                           '상의-소매길이': list_size_u3,
                           '상의-총장(앞)': list_size_u4,
                           '하의-총장(아웃심)': list_size_d1,
                           '하의-허리': list_size_d2,
                           '하의-엉덩이': list_size_d3,
                           '하의-허벅지': list_size_d4,
                           '하의-밑위': list_size_d5,
                           '하의-밑단': list_size_d6
                           }
                val['상품명'] = brand.iloc[brand_row_num]['제품명']
                val['구분(품번)'] = brand.iloc[brand_row_num]['상품코드']
                val['정상공급가'] = brand.iloc[brand_row_num]['가격']
                val['재고수량'] = brand.iloc[brand_row_num]['재고수량']
                val['원산지(제조국)(영문)'] = brand.iloc[brand_row_num]['원산지']
                val['소재'] = brand.iloc[brand_row_num]['소재']
                val['세탁방법'] = brand.iloc[brand_row_num]['세탁방법']
                val['제품 설명'] = brand.iloc[brand_row_num]['상품설명']
                val['색상(영문)'] = list_col[nol_col]
                val['사이즈'] = list_size[nol_size]
                val['상의-사이즈'] = list_size[nol_size]
                val['하의-사이즈'] = list_size[nol_size]
                val['재고수량'] = list_stock[k]

                data = data.append(val, ignore_index=True)

    # 참조정보가 없는 컬럼 빈칸 처리
    for column in china_columns:
        if (column in data.columns):
            continue
        else:
            data[column] = ""

    data = data[china_columns]
    return data


def df2excel(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['IMFORM']

    # 각 column의 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 8
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt+1).value = val
            row_cnt += 1
        col_cnt += 1

    wb.save(new_path)


def brand2china(file_path, upload_path):
    china_columns = ['구분(품번)', '상품명', '정상공급가', '할인율', '할인공급가', '색상(영문)', '사이즈', '사이즈(물산)', '재고수량', '원산지(제조국)(영문)', '카테고리(대분류)', '카테고리(중분류)', '카테고리(소분류)', '스타일+사이즈',
                     '상의-사이즈', '상의-어깨너비', '상의-가슴너비', '상의-소매길이', '상의-총장(앞)', '하의-사이즈', '하의-총장(아웃심)', '하의-허리', '하의-엉덩이', '하의-허벅지', '하의-밑위', '하의-밑단', '세탁방법', '품목 및 모델명', '소재', '제품 설명']

    brand_df = excel2df(file_path)
    del brand_df["이미지"]
    brand_df.drop(brand_df.index[1], inplace=True)

    brand_df = brand_df.reset_index()

    brand_df.drop(brand_df.columns[0], axis=1, inplace=True)

    china_df = generate_df(brand_df, china_columns)

    df2excel(china_df, UPLOAD_CHINA_FORM, upload_path)
