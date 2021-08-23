import pandas as pd
from path import UPLOAD_SEA_FORM
from openpyxl import load_workbook 


# def excel2df(file_path, sheet_name, a)
def excel2df(file_path):
    try:
        df = pd.read_excel(file_path)
        df = df.fillna("")
        df = df.applymap(lambda x: str(x))
        return df
    except:
        print("엑셀파일을 불러오는데 실패했습니다.")
        return 0


def generate_df(brand,order_columns):
    # 빈 데이터프레임 선언
    data = pd.DataFrame()
    product_code = pd.DataFrame()

    # 세부사이즈 컬럼 위치 저장
    for i in range(len(brand.columns)):
        if brand.columns[i] == '세부사이즈(상의)' :
            top_col_num = i
        if brand.columns[i] == '세부사이즈(하의)' :
            bottom_col_num = i

    # # 만약 컬럼명이 Unnamed로 정의되어 있다면, '그 전 컬럼의 이름 + 숫자'를 컬럼명으로 바꾸어줌.
    # for i in range(len(brand.columns)):
    #     if 'Unnamed:' in brand.columns[i]:
    #         list_num.append(str(i))
    #         brand.rename(columns={brand.columns[i] : brand.columns[i-1] + str(i)},inplace=True)

    for brand_row_num1 in range(len(brand.index)-1):
        # 사이즈 값이 없을 때 = one size, 색상 값이 없을 떄 = one color
        # 색,사이즈 개수가 2개 이상이면, row 추가해 각 컬럼에 색,사이즈 하나씩만 들어갈 수 있도록.
        # brand_row_num + 1 은 컬럼의 행이 2개이기 때문에 (상의, 하의부분)
        brand_row_num = brand_row_num1 + 1
        # 색상
        if brand.iloc[brand_row_num]["색상"] == "":
            list_col = ["ONE COLOR"]
        else:
            list_col = brand.iloc[brand_row_num]["색상"].split(',')
            list_col = [i.strip() for i in list_col] # 리스트 내에 있는 원소들의 양 옆 공백 없애기

        # 사이즈
        if brand.iloc[brand_row_num]["사이즈"] == "":
            list_size = ["ONE SIZE"]
        else:
            list_size = brand.iloc[brand_row_num]["사이즈"]
            # list_size = str(list_size)
            list_size = list_size.split(',')
            list_size = [i.strip() for i in list_size]

        if brand.iloc[brand_row_num]["무게(g)"] == "":
            list_weight = [""]
        else:
            list_weight = brand.iloc[brand_row_num]["무게(g)"]
            # list_weight = str(list_weight)
            list_weight = list_weight.split(',') # 이 부분 다시 확인.
            list_weight = [i.strip() for i in list_weight]
            list_weight = list(map(int, list_weight))
            for i in range(len(list_weight)):
                list_weight[i] = list_weight[i]/1000
        
        # print("가슴너비 : " + brand.iloc[0][top_col_num])
        
        # 분류에 따른 상세정보 입력
        classification = brand.iloc[brand_row_num]["분류"]
        classification = classification.split(',')
        classification = [i.strip() for i in classification]
        if "상의" in classification:
            des = "\n\n세부사이즈(상의)\n어깨너비 : " + brand.iloc[brand_row_num][top_col_num] + ", 가슴너비 : " + brand.iloc[brand_row_num][top_col_num+1] + ", 소매길이 : " + brand.iloc[brand_row_num][top_col_num+2] + ", 총장(앞) : " + brand.iloc[brand_row_num][top_col_num+3]
        if "하의" in classification:
            des = "\n\n세부사이즈(하의)\n총장(아웃심) : " + brand.iloc[brand_row_num][bottom_col_num] + ", 허리 : " + brand.iloc[brand_row_num][bottom_col_num + 1]+ ", 엉덩이 : " + brand.iloc[brand_row_num][bottom_col_num + 2] + ", 허벅지 : " + brand.iloc[brand_row_num][bottom_col_num+3]+ ", 밑위 : " + brand.iloc[brand_row_num][bottom_col_num + 4] + ", 밑단 : " + brand.iloc[brand_row_num][bottom_col_num+5]
        if "other" in classification:
            des = "\n\n" + brand.iloc[brand_row_num]["세부사이즈(other)"]

        # 색상x사이즈 개수 만큼 row 늘려 값 집어넣기
        for numOFlist_col in range(len(list_col)):
            for numOFlist_size in range(len(list_size)):
                val = {'Product Name' : brand.iloc[brand_row_num]["제품명"],
                        'Product Description' : brand.iloc[brand_row_num]["소재"] + str(des),
                        'Price' : brand.iloc[brand_row_num]["가격"],
                        'Stock' : brand.iloc[brand_row_num]["재고수량"],
                        'Weight' : list_weight[numOFlist_size],
                        # 'Variation Integration No.' : brand.iloc[brand_row_num]["상품코드"],
                        'Variation Name1' : "COLOR",
                        'Option for Variation 1' : list_col[numOFlist_col],
                        'Variation Name2' : "SIZE",
                        'Option for Variation 2' : list_size[numOFlist_size],
                        'Standard Express - Korea' : 'On'}

# 'Category','Maximum Purchase Quantity','Maximum Purchase Quantity - Start Date'
# ,'Maximum Purchase Quantity - Time Period (in Days)','Maximum Purchase Quantity - End Date','Parent SKU'
# ,'Image per Variation'
# ,'SKU','Cover image','Item image 1','Item image 2','Item image 3','Item image 4','Item image 5','Item image 6','Item image 7','Item image 8'
# ,'Length','Width','Height','Standard Express - Korea','Pre-order DTS'

                data = data.append(val, ignore_index=True)
        
        # print(data)

        # data["이미지"] = brand["상품코드"].apply(lambda x : [name for name in jpg_list if x in name])

    # 참조정보가 없는 컬럼 빈칸 처리
    for column in order_columns:
        if (column in data.columns):
            continue
        else:
            data[column] = ""

    data = data[order_columns]
    return data

def df2excel(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['Template']

    # 각 column의 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 6
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt).value = val
            row_cnt += 1
        col_cnt += 1

    wb.save(new_path)
    
def brand2SEA(file_path,upload_path):
    order_columns = ['Category','Product Name','Product Description','Maximum Purchase Quantity','Maximum Purchase Quantity - Start Date'
                    ,'Maximum Purchase Quantity - Time Period (in Days)','Maximum Purchase Quantity - End Date','Parent SKU','Variation Integration No.'
                    ,'Variation Name1','Option for Variation 1','Image per Variation','Variation Name2','Option for Variation 2','Price','Stock'
                    ,'SKU','Cover image','Item image 1','Item image 2','Item image 3','Item image 4','Item image 5','Item image 6','Item image 7','Item image 8'
                    ,'Weight','Length','Width','Height','Standard Express - Korea','Pre-order DTS']
                    
    brand_df = excel2df(file_path)
    # 정보 나타내는 행들 제거
    brand_df.drop(brand_df.index[1],inplace=True)
    # 인덱스 재정의
    brand_df = brand_df.reset_index()
    # reset_index 해주면서 생기는 index 컬럼 삭제
    brand_df.drop(brand_df.columns[0],axis=1,inplace=True)
    # print(type(brand_df.iloc[0]))  #<class 'pandas.core.series.Series'>
    sea_df = generate_df(brand_df,order_columns)
    df2excel(sea_df,UPLOAD_SEA_FORM,upload_path)