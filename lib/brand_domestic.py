import pandas as pd
from path import UPLOAD_DOMESTIC_FORM
from openpyxl import load_workbook 
pd.set_option("display.max_columns",20)

def excel2df(file_path):
    try:
        df = pd.read_excel(file_path)
        df = df.fillna("")
        return df
    except:
        print("엑셀파일을 불러오는데 실패했습니다.")
        return 0

def sum_list(x):
    try: 
        x_list = x.split(',')
        x_int_list=list(map(int, x_list))
        return sum(x_int_list)
    except:
        return x

def first_list(x):
    x_str = str(x)
    x_list = x_str.split(',')
    return x_list[0]

def option_price(color, size, price):
    Ncolor = len(color.split(','))
    Lcolor = (str(price)+",") * Ncolor
    Lcolor = Lcolor[:-1]
    
    Nsize = len(size.split(','))
    Lsize = (str(price)+",") * Nsize
    Lsize = Lsize[:-1]
    return Lcolor + "\n" + Lsize

def option_stock_num(color, stock):
    try:
        x_list = stock.split(',')
        x_int_list=list(map(int, x_list))
        Ncolor = len(color.split(','))
        bunch = int(len(x_int_list)/Ncolor)
        i = 0
        Lstock =[]
        for N in range(Ncolor,0,-1):
            Lstock.append(sum(x_int_list[i:i+bunch]))
            i += bunch
        Lstock = list(map(str, Lstock))
        return ','.join(Lstock)
    except:
        return stock


def generate_df(brand,order_columns):
    # 빈 데이터프레임 선언
    data = pd.DataFrame()

    # -------------------------------------통합데이터포맷 값 사용--------------------------------------------
    # [사용되지 않은 brand 컬럼들] : 상세정보(url), 세탁방법, 소재, 세부사이즈(other),세부사이즈(상의), 세부사이즈(하의),분류
    # 중간에 있는 컬럼명들은 서버에만 있는값.

    data["상품명"] = brand["제품명"].str[0:100]
    data["자체 상품코드"] = brand["상품코드"].str[0:50]
    # 카테고리ID
    data["요약설명"] = brand["상품설명"]
    data["재고수량"] = brand["재고수량"].apply(lambda x : sum_list(x))
    data["판매상태"] = data["재고수량"].apply(lambda x: "판매중" if x != 0 else "품절")
    data["상품상태"] = "신상품"
    data["판매가"] = brand["가격"]
    data["무게"] = brand["무게(g)"].apply(lambda x : first_list(x))
    # 정가
    data["재고사용"] = "Y"
    # SKU(재고번호)
    data["대표이미지파일명"] = brand["이미지"]
    data["상품 상세정보"] = brand["상세정보(html)"]
    data["세금"] = "과세상품"
    data["미성년자 구매"] = "Y"
    # 개인통관고유부호
    data["원산지"] = brand["원산지"]
    data["제조사"] = brand["제조사"].str.strip()
    data["브랜드"] = brand["브랜드명"].str.strip()
    # 배송방법, 택배 가능 여부, 퀵서비스 가능 여부, 방문 수령 가능 여부, 배송비 결제 방법, 배송비 유형,'기본배송비','조건부무료-상품판매가합계','무게별 차등 배송비(기본가격)','무게별 차등 배송비(무게)','무게별 차등 배송비(가격)','수량별 차등 배송비(타입)','수량별 차등 배송비(기본 가격)','수량별 차등 배송비(수량)','수량별 차등 배송비(가격)','수량별 차등 배송비(반복수량)','수량별 차등 배송비(반복가격)','구매금액별 차등 배송비(기본 가격)','구매금액 차등 배송비(구매금액)','구매금액별 차등 배송비(가격)',
    data["묶음배송 가능"] = "Y"
    data["별도설치비"] = "N"
    data["지역별 배송비 사용 여부"] = "A"
    # '지역별 배송비 유형','간편설정(제주도 추가 배송비)','간편설정(도서산간 추가 배송비)','우편번호 등록(시작구간)','우편번호 등록(종료구간)','우편번호 등록(추가배송비)','상품구매시 적립금 지급 값','상품구매시 적립금 지급 단위','할인적용대상','할인금액','할인금액 단위',
    data["옵션형태"] = "조합형"
    data["필수 옵션명"] = "색상\n사이즈"
    brand["색상"] = brand["색상"].apply(lambda x : x if x != "" else "ONE COLOR")
    brand["사이즈"] = brand["사이즈"].apply(lambda x : x if x != "" else "ONE SIZE")
    data["필수 옵션값"] = brand["색상"] + "\n" + brand["사이즈"]
    data["필수 옵션가"] = brand.apply(lambda x:option_price(x["색상"], x["사이즈"],x["가격"]), axis=1)
    data["옵션 재고수량"] = brand.apply(lambda x:option_stock_num(x["색상"],x["재고수량"]), axis=1)
    # '선택 옵션명','선택 옵션값','선택 옵션가','선택 옵션 재고수량','사용자 입력형 옵션','0원 선택옵션 최대 구매수량'
    data["재고소진후주문가능여부"] = "N"
    # "네이버/다음 쇼핑 노출용 상품명", '네이버 쇼핑 이벤트 문구','네이버 쇼핑 카테고리 ID','최소 구매수량','1회 구매시 최대 수량','1인 최대 구매수량'
    data["주문제작상품"] = "N"
    data["병행수입"] = "N"
    data["해외구매대행"] = "N"
    data["판매방식"] = "소매"
    data["다음 쇼핑하우 노출 설정"] = "N"
    data["네이버 쇼핑 노출 설정"] = "N"
    data["네이버 페이 구매가능 설정"] = "N"
    data["Facebook 다이내믹 광고 설정"] = "N"


    # 참조정보가 없는 컬럼 빈칸 처리
    for column in order_columns:
        if (column in data.columns):
            continue
        else:
            data[column] = ""

    data = data[order_columns]
    return data

def df2excel(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['ver.1.1']

    # 각 column의 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 4
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt).value = val
            row_cnt += 1
        col_cnt += 1

    wb.save(new_path)

def brand2domestic(file_path,upload_path):
    order_columns = ['상품명','자체 상품코드','카테고리ID','요약설명','판매상태','상품상태','판매가','무게','정가'
                        ,'재고사용','재고수량','SKU(재고번호)','대표 이미지 파일명','상품 상세정보','세금'
                        ,'미성년자 구매','개인통관고유부호','원산지','제조사','브랜드','배송방법','택배 가능 여부'
                        ,'퀵서비스 가능 여부','방문 수령 가능 여부','배송비 결제 방법','배송비 유형','기본배송비'
                        ,'조건부무료-상품판매가합계','무게별 차등 배송비(기본가격)','무게별 차등 배송비(무게)'
                        ,'무게별 차등 배송비(가격)','수량별 차등 배송비(타입)','수량별 차등 배송비(기본 가격)'
                        ,'수량별 차등 배송비(수량)','수량별 차등 배송비(가격)','수량별 차등 배송비(반복수량)'
                        ,'수량별 차등 배송비(반복가격)','구매금액별 차등 배송비(기본 가격)','구매금액 차등 배송비(구매금액)'
                        ,'구매금액별 차등 배송비(가격)','묶음배송 가능','별도설치비','지역별 배송비 사용 여부','지역별 배송비 유형'
                        ,'간편설정(제주도 추가 배송비)','간편설정(도서산간 추가 배송비)','우편번호 등록(시작구간)'
                        ,'우편번호 등록(종료구간)','우편번호 등록(추가배송비)','상품구매시 적립금 지급 값','상품구매시 적립금 지급 단위'
                        ,'할인적용대상','할인금액','할인금액 단위','옵션형태','필수 옵션명','필수 옵션값','필수 옵션가'
                        ,'옵션 재고수량','선택 옵션명','선택 옵션값','선택 옵션가','선택 옵션 재고수량','사용자 입력형 옵션'
                        ,'0원 선택옵션 최대 구매수량','재고소진후주문가능여부','네이버/다음 쇼핑 노출용 상품명','네이버 쇼핑 이벤트 문구'
                        ,'네이버 쇼핑 카테고리 ID','최소 구매수량','1회 구매시 최대 수량','1인 최대 구매수량'
                        ,'주문제작상품','병행수입','해외구매대행','판매방식','다음 쇼핑하우 노출 설정','네이버 쇼핑 노출 설정'
                        ,'네이버 페이 구매가능 설정','Facebook 다이내믹 광고 설정']
    
    brand_df = excel2df(file_path)
    # 정보 나타내는 행들 제거
    brand_df.drop([brand_df.index[0],brand_df.index[1]],inplace=True)

    domestic_df = generate_df(brand_df, order_columns)
    df2excel(domestic_df,UPLOAD_DOMESTIC_FORM,upload_path)
