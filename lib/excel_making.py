from pathlib import Path
import os
from openpyxl import load_workbook

FORM_PATH = Path(os.path.realpath(
    __file__)).parent.parent + '/transform/static/excel_form/excel_form.xlsx'
NEW_FILE_PATH = Path(os.path.realpath(
    __file__)).parent.parent + '/transform/static/files/new_excel.xlsx'
DOWN_FILE_PATH = Path(os.path.realpath(
    __file__)).parent.parent + '/tmp/data.xlsx'


excel_form = load_workbook(FORM_PATH, data_only=True)
order_excel = load_workbook(DOWN_FILE_PATH, data_only=True)
